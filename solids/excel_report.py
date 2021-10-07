from dagster import solid, SolidExecutionContext
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.fonts import Font
import pandas as pd


@solid(
    config_schema={
        'title': str,
        'subtitle': str
    }
)
def excel_report(context: SolidExecutionContext, df: pd.DataFrame) -> Workbook:
    wb = Workbook()
    ws = wb.worksheets[0]

    rows = dataframe_to_rows(df, index=False)

    # Write data into worksheet, starting on the 4th row to leave room for a title & subtitle
    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Add title & subtitle text
    ws.cell(1, 1, context.solid_config['title'])
    ws.cell(2, 1, context.solid_config['subtitle'])

    # Format it so it looks nice
    ws['A1'].style = 'Title'
    ws['A2'].style = 'Explanatory Text'

    # Bold the header row
    bold_font = Font(bold=True)
    for cell in ws["4:4"]:
        cell.font = bold_font

    # Set the columns to reasonable widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells[4:])
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

    return wb
